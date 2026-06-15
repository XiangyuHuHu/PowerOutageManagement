#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从洗煤厂设备新 CSV/XLSX 同步全部 Kep Tag 到 device_signal_points。"""
import argparse
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.database import get_db_cursor, init_database
from app.kep_opcua_address import (
    TAG_SUFFIX_TO_SIGNAL,
    build_nodeid,
    parse_equipment_parts,
    parse_tag_name,
    resolve_equipment_to_device,
)


def map_data_type(excel_type):
    raw = (excel_type or "").strip().lower()
    if raw == "boolean":
        return "bool"
    if raw in {"short", "int", "integer", "word", "ushort"}:
        return "int"
    if raw in {"float", "double", "real"}:
        return "float"
    return "string"


def load_csv_tags(path: Path):
    for enc in ("utf-8-sig", "gbk", "gb18030"):
        try:
            with path.open(encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                col = reader.fieldnames[0] if reader.fieldnames else "Tag Name"
                for row in reader:
                    tag = (row.get(col) or row.get("Tag Name") or "").strip()
                    parsed = parse_tag_name(tag)
                    if parsed:
                        yield parsed[0], parsed[1], tag
            return
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"无法解码 CSV: {path}")


def load_xlsx_tags(path: Path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        tag = str(row[0]).strip()
        if tag in {"Tag Name", "Name"} or "." not in tag:
            continue
        parsed = parse_tag_name(tag)
        if not parsed:
            continue
        equipment, suffix = parsed
        yield {
            "equipment": equipment,
            "suffix": suffix,
            "signal_type": TAG_SUFFIX_TO_SIGNAL[suffix],
            "full_tag": tag,
            "plc_address": (row[1] or "").strip() if len(row) > 1 else "",
            "data_type": map_data_type(row[2] if len(row) > 2 else ""),
            "description": (row[15] or "").strip() if len(row) > 15 else "",
        }
    wb.close()


def collect_tag_items(source_path: Path):
    if source_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return list(load_xlsx_tags(source_path))
    items = []
    for equipment, suffix, full_tag in load_csv_tags(source_path):
        items.append(
            {
                "equipment": equipment,
                "suffix": suffix,
                "signal_type": TAG_SUFFIX_TO_SIGNAL[suffix],
                "full_tag": full_tag,
                "plc_address": "",
                "data_type": "bool" if TAG_SUFFIX_TO_SIGNAL[suffix] == "power_feedback" else "int",
                "description": full_tag,
            }
        )
    return items


def ensure_devices_from_tags(tag_items):
    equipment_set = sorted({item["equipment"] for item in tag_items if item.get("equipment")})
    if not equipment_set:
        return 0
    created = 0
    with get_db_cursor() as cursor:
        for equipment in equipment_set:
            device_id, device_name = parse_equipment_parts(equipment)
            if not device_id:
                continue
            cursor.execute(
                """
                INSERT INTO devices (device_id, device_name, power_room, cabinet, area_code, line_name, sort_order, is_active)
                VALUES (%s, %s, '待归类', '', 'kep-import', '', 9000, TRUE)
                ON DUPLICATE KEY UPDATE
                    device_name = IF(device_name IS NULL OR device_name = '', VALUES(device_name), device_name),
                    is_active = TRUE
                """,
                (device_id, device_name),
            )
            if cursor.rowcount == 1:
                created += 1
    return created


def upsert_signal_rows(rows):
    sql = """
        INSERT INTO device_signal_points
        (device_id, signal_type, signal_name, signal_address, data_type, source_system, source_sheet, description, is_active)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE)
        ON DUPLICATE KEY UPDATE
            signal_address = VALUES(signal_address),
            data_type = VALUES(data_type),
            source_system = VALUES(source_system),
            source_sheet = VALUES(source_sheet),
            description = VALUES(description),
            is_active = TRUE
    """
    with get_db_cursor() as cursor:
        cursor.executemany(sql, rows)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("source", nargs="?", help="CSV 或 XLSX 路径")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--sql-out", type=str, help="仅根据台账种子生成 SQL 文件，不连数据库")
    p.add_argument("--no-ensure-devices", action="store_true", help="不自动补全缺失设备台账")
    p.add_argument(
        "--md-out",
        nargs="?",
        const=str(ROOT / "DEVICE_SIGNAL_MATCH_LOG.md"),
        default=None,
        help="导入后刷新匹配记录 Markdown（默认 DEVICE_SIGNAL_MATCH_LOG.md）",
    )
    args = p.parse_args()
    source_path = Path(args.source) if args.source else Path(r"d:\金海泽地\金正泰\洗煤厂设备新0601.xlsx")
    if not source_path.is_file():
        fallback = ROOT / "data" / "洗煤厂设备新0601.csv"
        source_path = fallback if fallback.is_file() else source_path
    if not source_path.is_file():
        print("文件不存在:", source_path)
        return 1

    tag_items = collect_tag_items(source_path)

    if args.sql_out:
        from app.database import REAL_DEVICE_SEED

        devices = [{"device_id": row[0], "device_name": row[1]} for row in REAL_DEVICE_SEED]
        for item in tag_items:
            did, dname = parse_equipment_parts(item["equipment"])
            if did and did not in {d["device_id"] for d in devices}:
                devices.append({"device_id": did, "device_name": dname})
    else:
        init_database()
        if not args.no_ensure_devices:
            created = ensure_devices_from_tags(tag_items)
            print("自动补全设备台账:", created, "台（新建）")
        with get_db_cursor() as cursor:
            cursor.execute("SELECT device_id, device_name FROM devices WHERE is_active = TRUE")
            devices = cursor.fetchall()

    db_rows = []
    missing = []
    seen = set()
    sheet = "xlsx0601" if source_path.suffix.lower() in {".xlsx", ".xlsm"} else "csv0601"

    for item in tag_items:
        dev = resolve_equipment_to_device(item["equipment"], devices)
        if not dev:
            missing.append(item["full_tag"])
            continue
        key = (dev["device_id"], item["signal_type"], item["suffix"])
        if key in seen:
            continue
        seen.add(key)
        desc = item["description"] or ("Kep: " + item["full_tag"])
        if item["plc_address"]:
            desc = desc + " | PLC: " + item["plc_address"]
        db_rows.append(
            (
                dev["device_id"],
                item["signal_type"],
                item["suffix"],
                build_nodeid(item["full_tag"]),
                item["data_type"],
                "kepserver",
                sheet,
                desc[:255],
            )
        )

    print("识别点位:", len(seen), "匹配写入:", len(db_rows), "未匹配:", len(missing))
    if missing[:5]:
        print("未匹配示例:", missing[:5])

    if args.dry_run:
        for u in db_rows[:5]:
            print("示例", u)
        return 0

    if args.sql_out:
        out = Path(args.sql_out)
        lines = [
            "-- generated from Kep xlsx/csv; run in MySQL power_control",
            "SET NAMES utf8mb4;",
        ]
        for row in db_rows:
            device_id, st, suffix, nodeid, dtype, src, sheet, desc = row
            esc = lambda s: (s or "").replace("\\", "\\\\").replace("'", "''")
            lines.append(
                "INSERT INTO device_signal_points "
                "(device_id, signal_type, signal_name, signal_address, data_type, source_system, source_sheet, description, is_active) "
                f"VALUES ('{esc(device_id)}', '{esc(st)}', '{esc(suffix)}', '{esc(nodeid)}', '{esc(dtype)}', '{esc(src)}', '{esc(sheet)}', '{esc(desc)}', TRUE) "
                "ON DUPLICATE KEY UPDATE signal_address=VALUES(signal_address), data_type=VALUES(data_type), "
                "source_system=VALUES(source_system), source_sheet=VALUES(source_sheet), description=VALUES(description), is_active=TRUE;"
            )
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print("已生成 SQL:", out, "条数:", len(db_rows))
        return 0

    upsert_signal_rows(db_rows)
    print("已写入 device_signal_points")
    if args.md_out:
        import sys
        from tools.refresh_device_signal_match_log import main as refresh_log

        argv_save = sys.argv
        sys.argv = ["refresh_device_signal_match_log", "--md-out", args.md_out]
        try:
            refresh_log()
        finally:
            sys.argv = argv_save
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
