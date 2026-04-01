import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def looks_like_device_id(value: str) -> bool:
    if not value:
        return False
    allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_-/#")
    return all(ch in allowed for ch in value)


def export_rows(workbook_path: Path):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    devices = []
    signals = []

    for ws in wb.worksheets:
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [text(v) for v in row[:12]]
            while len(values) < 12:
                values.append("")
            col1, col2, col3, col4, col5, _, col7, col8, _, col10, _, col12 = values

            device_id = ""
            device_name = ""
            if looks_like_device_id(col2) and col3:
                device_id, device_name = col2, col3
            elif looks_like_device_id(col3) and col4:
                device_id, device_name = col3, col4

            if device_id:
                display_name = col5 or col4 or col1 or f"{device_id} {device_name}"
                devices.append(
                    {
                        "source_sheet": ws.title,
                        "row_number": row_number,
                        "sequence": col1,
                        "device_id": device_id,
                        "device_name": device_name,
                        "display_name": display_name,
                        "related_device_name": col7,
                        "related_device_id": col8,
                        "db_address": col10,
                        "remote_local_input": col12,
                    }
                )
                if col12:
                    signals.append(
                        {
                            "device_id": device_id,
                            "signal_type": "remote_local",
                            "signal_name": "remote_local",
                            "signal_address": col12,
                            "data_type": "bool",
                            "source_system": "plc",
                            "source_sheet": ws.title,
                            "description": f"linked register: {col10}" if col10 else "",
                            "is_active": True,
                        }
                    )

            if "YCFHZ" in col4 and col5:
                signals.append(
                    {
                        "device_id": col5.split("_")[0],
                        "signal_type": "remote_switch_binding",
                        "signal_name": "remote_switch_binding",
                        "signal_address": col4,
                        "data_type": "binding",
                        "source_system": "plc",
                        "source_sheet": ws.title,
                        "description": col5,
                        "is_active": True,
                    }
                )

    seen_devices = set()
    unique_devices = []
    for row in devices:
        key = (row["device_id"], row["source_sheet"], row["row_number"])
        if key not in seen_devices:
            seen_devices.add(key)
            unique_devices.append(row)

    seen_signals = set()
    unique_signals = []
    for row in signals:
        key = (row["device_id"], row["signal_type"], row["signal_name"], row["signal_address"])
        if key not in seen_signals:
            seen_signals.add(key)
            unique_signals.append(row)

    summary = {
        "workbook": str(workbook_path),
        "device_rows": len(unique_devices),
        "signal_rows": len(unique_signals),
        "unique_device_ids": len({row["device_id"] for row in unique_devices}),
        "signal_types": {},
    }
    for row in unique_signals:
        summary["signal_types"][row["signal_type"]] = summary["signal_types"].get(row["signal_type"], 0) + 1

    return unique_devices, unique_signals, summary


def main():
    parser = argparse.ArgumentParser(description="Export normalized signal ledger rows from workbook.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--devices-output", required=True)
    parser.add_argument("--signals-output", required=True)
    parser.add_argument("--summary-output", required=True)
    args = parser.parse_args()

    devices, signals, summary = export_rows(Path(args.input))
    Path(args.devices_output).write_text(json.dumps(devices, ensure_ascii=False, indent=2), encoding="utf-8")
    Path(args.signals_output).write_text(json.dumps(signals, ensure_ascii=False, indent=2), encoding="utf-8")
    Path(args.summary_output).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"devices={len(devices)} signals={len(signals)}")


if __name__ == "__main__":
    main()
